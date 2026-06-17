"""XLSX 样式感知抽取器：每个 sheet 产出一张 StyledGrid（L1 确定性，零 LLM）。

职责（PRD user stories 1–5）：
    - theme 色 + tint 解析为真实 'RRGGBB'（不能把 theme 引用直接当颜色）。
    - 合并单元格展开：锚点格标 is_merged_origin + merge_span，承载多级表头语义。
    - 数字格式保留：百分比 / 千分位 / 货币 / 小数位写入 StyledCell.number_format。
    - 条件格式区域检测：落在 CF 规则范围内的格打 cf_affected=True，并在 grid 级
      追加一条 warning（颜色来源不可靠，下游颜色语义跳过）。
    - source_file_hash 写入每张 grid，作为版本与审计血缘根。
"""

import colorsys
import hashlib
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from ragspine.extraction.ir import StyledCell, StyledGrid

_A_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"

# Excel theme 调色板顺序（clrScheme 子元素顺序，index 与单元格 theme 引用对齐）。
_THEME_SLOTS = (
    "dk1", "lt1", "dk2", "lt2",
    "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
    "hlink", "folHlink",
)


def compute_file_hash(path: str | Path) -> str:
    """计算源文件内容 hash（十六进制串），作为版本血缘标识。"""
    digest = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def resolve_theme_color(theme_index: int, tint: float, theme_rgbs: list[str]) -> str:
    """把 (theme 调色板索引, tint) 解析为 'RRGGBB' 大写十六进制。

    tint > 0 向白调亮、tint < 0 向黑调暗，算法遵循 OOXML 规范。
    theme_rgbs 为从 workbook 主题 XML 解析出的基色表。
    """
    base = theme_rgbs[theme_index].upper()
    if not tint:
        return base
    r = int(base[0:2], 16) / 255.0
    g = int(base[2:4], 16) / 255.0
    b = int(base[4:6], 16) / 255.0
    h, lum, s = colorsys.rgb_to_hls(r, g, b)
    if tint < 0:
        lum = lum * (1.0 + tint)
    else:
        lum = lum * (1.0 - tint) + tint
    r2, g2, b2 = colorsys.hls_to_rgb(h, lum, s)
    return "{:02X}{:02X}{:02X}".format(
        round(r2 * 255), round(g2 * 255), round(b2 * 255)
    )


def _parse_theme_rgbs(wb) -> list[str]:
    """从 workbook 主题 XML 解析有序基色表（与单元格 theme index 对齐）。"""
    rgbs = ["FFFFFF"] * len(_THEME_SLOTS)
    raw = getattr(wb, "loaded_theme", None)
    if not raw:
        return rgbs
    root = ET.fromstring(raw)
    scheme = root.find(f".//{_A_NS}clrScheme")
    if scheme is None:
        return rgbs
    for idx, slot in enumerate(_THEME_SLOTS):
        node = scheme.find(f"{_A_NS}{slot}")
        if node is None:
            continue
        srgb = node.find(f"{_A_NS}srgbClr")
        if srgb is not None and srgb.get("val"):
            rgbs[idx] = srgb.get("val").upper()
            continue
        sys_clr = node.find(f"{_A_NS}sysClr")
        if sys_clr is not None and sys_clr.get("lastClr"):
            rgbs[idx] = sys_clr.get("lastClr").upper()
    return rgbs


def _normalize_rgb(argb: str) -> str:
    """把 openpyxl 'AARRGGBB'/'RRGGBB' 规范成 6 位大写 'RRGGBB'。"""
    argb = argb.upper()
    if len(argb) == 8:
        argb = argb[2:]
    return argb


def _resolve_fill_rgb(cell, theme_rgbs: list[str]) -> str | None:
    """解析单元格填充色为真实 'RRGGBB'；无填充返回 None。"""
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    if fg.type == "rgb" and fg.rgb:
        rgb = _normalize_rgb(str(fg.rgb))
        # openpyxl 默认值 '00000000' 视为无填充
        if rgb == "000000" and str(fg.rgb).upper() == "00000000":
            return None
        return rgb
    if fg.type == "theme":
        try:
            return resolve_theme_color(int(fg.theme), float(fg.tint or 0.0), theme_rgbs)
        except (IndexError, ValueError, TypeError):
            return None
    return None


def _build_grid(ws, doc_id: str, file_hash: str, theme_rgbs: list[str]) -> StyledGrid:
    grid = StyledGrid(
        sheet=ws.title,
        source_doc_id=doc_id,
        source_file_hash=file_hash,
        n_rows=ws.max_row,
        n_cols=ws.max_column,
    )

    # 合并区域：锚点 ref -> (n_rows, n_cols)
    merge_origins: dict[str, tuple[int, int]] = {}
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        origin = f"{get_column_letter(min_col)}{min_row}"
        merge_origins[origin] = (max_row - min_row + 1, max_col - min_col + 1)

    # 条件格式区域：收集受影响的单元格坐标
    cf_refs: set[str] = set()
    cf_ranges: list[str] = []
    for cf in ws.conditional_formatting:
        sqref = str(cf.sqref)
        cf_ranges.append(sqref)
        for token in sqref.split():
            min_col, min_row, max_col, max_row = range_boundaries(token)
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    cf_refs.add(f"{get_column_letter(c)}{r}")
    if cf_ranges:
        grid.add_warning(
            "检测到条件格式区域 "
            + ", ".join(cf_ranges)
            + "：颜色来源不可靠，已跳过颜色语义"
        )

    for row in ws.iter_rows():
        for cell in row:
            ref = cell.coordinate
            cf_affected = ref in cf_refs
            resolved_rgb = None if cf_affected else _resolve_fill_rgb(cell, theme_rgbs)
            span = merge_origins.get(ref)
            value = cell.value

            # 稀疏映射：只保留有值 / 有填充 / 有合并 / 落在 CF 区域的格
            if value is None and resolved_rgb is None and span is None and not cf_affected:
                continue

            grid.cells[ref] = StyledCell(
                value=value,
                cell_ref=ref,
                resolved_rgb=resolved_rgb,
                number_format=cell.number_format or "General",
                bold=bool(cell.font and cell.font.bold),
                is_merged_origin=span is not None,
                merge_span=span,
                cf_affected=cf_affected,
            )

    return grid


def extract_grids(path: str | Path) -> list[StyledGrid]:
    """抽取一个 xlsx 的全部 StyledGrid（每个 worksheet 一张）。"""
    path = Path(path)
    doc_id = path.name
    file_hash = compute_file_hash(path)
    wb = load_workbook(str(path), data_only=True)
    theme_rgbs = _parse_theme_rgbs(wb)
    return [_build_grid(ws, doc_id, file_hash, theme_rgbs) for ws in wb.worksheets]
