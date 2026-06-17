"""XLSX 确定性抽取：5-yr summary 表直接按 schema 映射，零幻觉。

约定（5yr summary 表）：
    - A 列：指标名（A2 起），经 glossary 归一为 metric_code。
    - 第 1 行：期间表头（B1 起），经 glossary 归一为 (period_type, period)。
    - 实体：优先取 B1 同行旁的 A1 单元格约定，否则用 sheet 名解析。
    - source_locator 用真实单元格坐标，如 'sheet=5yr!C4'。
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ragspine.common.glossary import (
    geography_for_entity,
    normalize_entity,
    normalize_metric,
    normalize_period,
    unit_for_metric,
)
from ragspine.storage.fact_store import Fact


def _coerce_number(value: object) -> float | None:
    """单元格值转数值；非数值返回 None。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "").replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def extract_facts(path: str | Path) -> tuple[list[Fact], list[str]]:
    """抽取一个 xlsx 的全部 Fact。返回 (facts, warnings)。"""
    path = Path(path)
    doc_id = path.name
    wb = load_workbook(str(path), data_only=True)
    facts: list[Fact] = []
    warnings: list[str] = []

    for ws in wb.worksheets:
        # 实体约定：A1 单元格写实体名；否则退回 sheet 名
        entity = normalize_entity(ws["A1"].value) or normalize_entity(ws.title)
        if entity is None:
            warnings.append(f"sheet={ws.title}: A1/sheet 名无法解析实体，跳过该表")
            continue
        geography = geography_for_entity(entity) or "UNKNOWN"

        # 期间表头：第 1 行 B 列起
        periods: dict[int, tuple[tuple[str, str], str]] = {}
        for col in range(2, ws.max_column + 1):
            raw = ws.cell(row=1, column=col).value
            parsed = normalize_period(str(raw)) if raw is not None else None
            if parsed is None:
                if raw is not None:
                    warnings.append(
                        f"sheet={ws.title}!{get_column_letter(col)}1: 无法识别期间 '{raw}'，跳过该列"
                    )
                continue
            periods[col] = (parsed, str(raw).strip())

        # 指标行：第 2 行起 A 列
        for row in range(2, ws.max_row + 1):
            metric_label = ws.cell(row=row, column=1).value
            metric_code = normalize_metric(str(metric_label)) if metric_label is not None else None
            if metric_code is None:
                if metric_label is not None:
                    warnings.append(
                        f"sheet={ws.title}!A{row}: 无法识别指标 '{metric_label}'，跳过该行"
                    )
                continue
            unit = unit_for_metric(metric_code) or "USD_M"
            for col, ((period_type, period), _period_raw) in periods.items():
                value = _coerce_number(ws.cell(row=row, column=col).value)
                if value is None:
                    continue
                coord = f"{get_column_letter(col)}{row}"
                facts.append(
                    Fact(
                        metric_code=metric_code,
                        entity=entity,
                        geography=geography,
                        channel="TOTAL",
                        period_type=period_type,
                        period=period,
                        value=value,
                        unit=unit,
                        source_doc_id=doc_id,
                        source_locator=f"sheet={ws.title}!{coord}",
                    )
                )

    return facts, warnings
