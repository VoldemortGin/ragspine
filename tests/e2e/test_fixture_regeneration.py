"""合成 fixture 一键再生测试（PRD user story #20）。

story #20: As 数据工程师, I want 合成 golden 文档可一键再生,
so that 测试与评测不依赖任何真实敏感数据。

scripts/make_fixtures_excel.py 已是真实实现（红色阶段交付），本模块验证其
**既有真实行为**——删除/连续生成都得到与 ground truth 等价的确定性产物，
因此这些用例**预期 PASS**（对已实现能力的覆盖补缺，不是红测）。

只验证外部行为：用 openpyxl 把再生文件读回，断言关键格的值/填充色/数字格式/
合并范围/条件格式区域与 ground truth 一致；连续两次再生内容等价；ground truth
的逐格清单与再生文件保持一致性。

安全约束：generator 写死目标路径（XLSX_PATH/GT_PATH 指向 data/fixtures/），
无法重定向。因此每个会改动 data/fixtures/ 的用例都**先把真实 fixture 备份到 tmp，
再在 finally 中无条件恢复**——无论断言成败，data/fixtures/ 都回到测试前状态。
"""

import json
import os
import shutil

import pytest
import rootutils

ROOT_DIR = rootutils.setup_root(os.getcwd(), indicator=".project-root", pythonpath=True)

from openpyxl import load_workbook

from ragspine.fixtures.excel import (
    GT_PATH,
    XLSX_PATH,
    main as make_excel_fixtures,
)


# ---------------------------------------------------------------------------
# 安全网：备份真实 fixture -> 运行 generator -> finally 恢复
# ---------------------------------------------------------------------------

@pytest.fixture
def fixture_sandbox(tmp_path):
    """在 tmp 备份 data/fixtures/ 下两个真实产物，测试后无条件恢复。

    yield 一个回调：调用即（在备份就绪后）运行 generator，把真实路径重建。
    无论用例断言是否失败，teardown 都用备份覆盖回 XLSX_PATH/GT_PATH，
    确保 data/fixtures/ 不被本测试弄坏。
    """
    backup_xlsx = tmp_path / "backup_fixture.xlsx"
    backup_gt = tmp_path / "backup_ground_truth.json"
    had_xlsx = XLSX_PATH.exists()
    had_gt = GT_PATH.exists()
    if had_xlsx:
        shutil.copy2(XLSX_PATH, backup_xlsx)
    if had_gt:
        shutil.copy2(GT_PATH, backup_gt)

    try:
        yield
    finally:
        # 无条件恢复到测试前状态
        if had_xlsx:
            shutil.copy2(backup_xlsx, XLSX_PATH)
        elif XLSX_PATH.exists():
            XLSX_PATH.unlink()
        if had_gt:
            shutil.copy2(backup_gt, GT_PATH)
        elif GT_PATH.exists():
            GT_PATH.unlink()


# ---------------------------------------------------------------------------
# 读回辅助：把 openpyxl 的填充归一为可比较表示
# ---------------------------------------------------------------------------

def _effective_fill(cell):
    """返回再生文件中某格填充的可比较表示。

    - 无真实填充（patternType is None）-> None
    - theme 填充 -> ('theme', theme_index, round(tint, 6))
    - 普通 rgb 填充 -> 末 6 位大写十六进制 'RRGGBB'
    """
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    fg = fill.fgColor
    if fg.type == "theme":
        return ("theme", fg.theme, round(float(fg.tint), 6))
    if fg.type == "rgb":
        return str(fg.rgb)[-6:].upper()
    return None


def _expected_fill(truth_cell: dict, palette: dict):
    """把 ground truth 的 resolved_rgb 翻译成再生文件应有的填充表示。

    theme 解析出的 resolved_rgb（95B3D7）在 xlsx 里仍以 theme 引用存储，
    故对该值期望的是 ('theme', accent1_index, accent1_tint)。
    """
    rgb = truth_cell.get("resolved_rgb")
    if rgb is None:
        return None
    if rgb == palette["accent1_resolved_rgb"]:
        return ("theme", palette["accent1_index"], round(float(palette["accent1_tint"]), 6))
    return rgb.upper()


def _read_workbook_state(xlsx_path) -> dict:
    """把整个工作簿读成可比较的纯数据结构（值 + 填充 + 数字格式 + 合并 + CF）。"""
    wb = load_workbook(str(xlsx_path))
    state: dict = {"sheets": {}}
    for ws in wb.worksheets:
        cells: dict[str, dict] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None and (cell.fill is None or cell.fill.patternType is None):
                    continue
                cells[cell.coordinate] = {
                    "value": cell.value,
                    "fill": _effective_fill(cell),
                    "number_format": cell.number_format,
                }
        merges = sorted(str(r) for r in ws.merged_cells.ranges)
        cf_ranges = sorted(str(cf.sqref) for cf in ws.conditional_formatting)
        state["sheets"][ws.title] = {
            "cells": cells,
            "merges": merges,
            "cf_ranges": cf_ranges,
        }
    return state


# ===========================================================================
# story #20 —— 用例 1：删除后一键再生，且与 ground truth 一致（确定性再生）
# ===========================================================================

def test_regenerate_after_deletion_rebuilds_files(fixture_sandbox):
    """story #20 —— 删除两个产物后调用生成器，文件被重建且非空。"""
    XLSX_PATH.unlink()
    GT_PATH.unlink()
    assert not XLSX_PATH.exists()
    assert not GT_PATH.exists()

    make_excel_fixtures()

    assert XLSX_PATH.exists(), "xlsx 未被重建"
    assert GT_PATH.exists(), "ground truth 未被重建"
    assert XLSX_PATH.stat().st_size > 0
    assert GT_PATH.stat().st_size > 0
    # ground truth 是合法 JSON 且结构完整
    gt = json.loads(GT_PATH.read_text(encoding="utf-8"))
    assert {"file", "theme_palette", "sheets", "cells"} <= set(gt)
    assert gt["cells"], "再生 ground truth 无逐格真值"


def test_regenerated_cell_values_match_ground_truth(fixture_sandbox):
    """story #20 —— 删除后再生，openpyxl 读回的逐格值与 ground truth 完全一致。"""
    XLSX_PATH.unlink()
    GT_PATH.unlink()
    make_excel_fixtures()

    gt = json.loads(GT_PATH.read_text(encoding="utf-8"))
    wb = load_workbook(str(XLSX_PATH))
    for t in gt["cells"]:
        ws = wb[t["sheet"]]
        cell = ws[t["cell_ref"]]
        assert cell.value == t["value"], f"{t['sheet']}!{t['cell_ref']} 值不符"


def test_regenerated_cell_styles_match_ground_truth(fixture_sandbox):
    """story #20 —— 再生文件的填充色/数字格式逐格与 ground truth 一致。

    theme+tint 格以 theme 引用比对（resolved_rgb 95B3D7 -> ('theme',4,0.4)）；
    普通 RGB 格比末 6 位；resolved_rgb=None 的格须无真实填充。
    """
    XLSX_PATH.unlink()
    GT_PATH.unlink()
    make_excel_fixtures()

    gt = json.loads(GT_PATH.read_text(encoding="utf-8"))
    palette = gt["theme_palette"]
    wb = load_workbook(str(XLSX_PATH))
    for t in gt["cells"]:
        ws = wb[t["sheet"]]
        cell = ws[t["cell_ref"]]
        ref = f"{t['sheet']}!{t['cell_ref']}"
        assert _effective_fill(cell) == _expected_fill(t, palette), f"{ref} 填充不符"
        assert cell.number_format == t["number_format"], f"{ref} 数字格式不符"


def test_regenerated_merges_and_cf_match_ground_truth(fixture_sandbox):
    """story #20 —— 再生文件的合并范围与条件格式区域与 ground truth 元数据一致。"""
    XLSX_PATH.unlink()
    GT_PATH.unlink()
    make_excel_fixtures()

    gt = json.loads(GT_PATH.read_text(encoding="utf-8"))
    wb = load_workbook(str(XLSX_PATH))

    # 合并表头范围
    merged_meta = gt["sheets"]["MergedHeader"]["merges_expect"]
    expected_merges = {m["range"] for m in merged_meta}
    actual_merges = {str(r) for r in wb["MergedHeader"].merged_cells.ranges}
    assert expected_merges <= actual_merges, (expected_merges, actual_merges)

    # 条件格式区域
    cf_meta = gt["sheets"]["CondFormat"]["cf_ranges_expect"]
    actual_cf = {str(cf.sqref) for cf in wb["CondFormat"].conditional_formatting}
    assert set(cf_meta) <= actual_cf, (cf_meta, actual_cf)


def test_regenerated_legend_text_matches_ground_truth(fixture_sandbox):
    """story #20 —— 再生文件图例区文字与色块填充与 ground truth 一致。"""
    XLSX_PATH.unlink()
    GT_PATH.unlink()
    make_excel_fixtures()

    gt = json.loads(GT_PATH.read_text(encoding="utf-8"))
    ws = load_workbook(str(XLSX_PATH))["HK_Performance"]
    for item in gt["sheets"]["HK_Performance"]["legend_expect"]:
        assert ws[item["text_ref"]].value == item["meaning"]
        swatch = ws[item["swatch_ref"]]
        assert _effective_fill(swatch) == item["rgb"].upper()


# ===========================================================================
# story #20 —— 用例 2：连续两次再生内容等价（值与样式层面，非字节级）
# ===========================================================================

def test_two_regenerations_workbook_state_equivalent(fixture_sandbox):
    """story #20 —— 连续生成两次，工作簿的值/填充/格式/合并/CF 全等价。"""
    make_excel_fixtures()
    state_a = _read_workbook_state(XLSX_PATH)

    make_excel_fixtures()
    state_b = _read_workbook_state(XLSX_PATH)

    assert state_a == state_b


def test_two_regenerations_ground_truth_equivalent(fixture_sandbox):
    """story #20 —— 连续生成两次，ground truth JSON 内容完全等价。"""
    make_excel_fixtures()
    gt_a = json.loads(GT_PATH.read_text(encoding="utf-8"))

    make_excel_fixtures()
    gt_b = json.loads(GT_PATH.read_text(encoding="utf-8"))

    assert gt_a == gt_b


# ===========================================================================
# story #20 —— 用例 3：ground truth 逐格清单 与 再生文件 一致性抽查
# ===========================================================================

def test_ground_truth_cell_list_consistent_with_regenerated_file(fixture_sandbox):
    """story #20 —— ground truth 列出的每个格，在再生文件里都真实存在且自洽。

    抽查一致性：每条逐格真值的 (sheet, cell_ref) 在再生文件里可定位，
    且该格至少承载真值断言的内容之一（值 / 填充），不是凭空记录。
    """
    make_excel_fixtures()
    gt = json.loads(GT_PATH.read_text(encoding="utf-8"))
    palette = gt["theme_palette"]
    wb = load_workbook(str(XLSX_PATH))

    assert wb.sheetnames == list(gt["sheets"].keys()), (
        wb.sheetnames, list(gt["sheets"].keys()))

    for t in gt["cells"]:
        ref = f"{t['sheet']}!{t['cell_ref']}"
        assert t["sheet"] in wb.sheetnames, f"{ref} 的 sheet 不存在"
        cell = wb[t["sheet"]][t["cell_ref"]]
        # 真值记录的格不能在文件里既无值又无填充（即必须是真实承载内容的格）
        has_value = cell.value is not None
        has_fill = _effective_fill(cell) is not None
        assert has_value or has_fill, f"{ref} 在再生文件里既无值也无填充"
        # 抽查值与填充自洽
        assert cell.value == t["value"], f"{ref} 值与真值不符"
        assert _effective_fill(cell) == _expected_fill(t, palette), f"{ref} 填充与真值不符"


def test_ground_truth_covers_all_styled_data_cells(fixture_sandbox):
    """story #20 —— 再生文件里所有承载值的格都被 ground truth 逐格清单覆盖。

    反向抽查：避免 ground truth 漏记某些值格导致评测覆盖出现盲区。
    （图例色块格 value=None 但有填充，已在 ground truth 中单列；此处只查有值格。）
    """
    make_excel_fixtures()
    gt = json.loads(GT_PATH.read_text(encoding="utf-8"))
    gt_keys = {f"{c['sheet']}!{c['cell_ref']}" for c in gt["cells"]}

    wb = load_workbook(str(XLSX_PATH))
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                key = f"{ws.title}!{cell.coordinate}"
                assert key in gt_keys, f"再生文件有值格 {key} 未被 ground truth 记录"
