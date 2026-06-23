"""生成一期（Excel 线）合成 fixture + 逐格 ground truth（确定性、硬编码）。

覆盖 PRD 一期全部刁钻形态（user stories 1–7）：
    - theme 色 + tint 填充 与 普通 RGB 填充；同色组跨行列分布（颜色=属性语义）。
    - 一个图例区（色块格 + 文字「黄色=新产品线」）。
    - 三级合并表头 + 一张转置表（指标在列、期间在行）。
    - 百分比 / 千分位 / 货币 格式数字格。
    - 一个带条件格式规则（色阶）的区域。
    - 正常的指标×期间数据区（复用 glossary 的 REVENUE/NEWSALES/PROFIT/ROE 与 ACME_HK 受控词）。

产出：
    data/fixtures/excel_styled_fixture.xlsx
    data/fixtures/fixtures_ground_truth.json   —— 逐格真值（值/RGB/合并/格式/tags/
                                                  图例映射期望/CF 区域清单）。

末尾用 openpyxl 重新读回自校验（值与样式写入成功），确保 fixture 本身可靠。
从项目根目录运行：.venv/bin/python scripts/make_fixtures_excel.py
"""

import colorsys
import json
from pathlib import Path

import rootutils
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

ROOT_DIR = rootutils.find_root(Path(__file__), indicator=".project-root")

OUT_DIR = ROOT_DIR / "data" / "fixtures"
XLSX_PATH = OUT_DIR / "excel_styled_fixture.xlsx"
GT_PATH = OUT_DIR / "fixtures_ground_truth.json"

# --- 默认主题调色板（openpyxl 默认 theme，accent 段无 dk/lt 交换）-----------
# 与 src 抽取器将解析的 theme 基色一致；index 4 = accent1。
THEME_ACCENT1_INDEX = 4
THEME_ACCENT1_RGB = "4F81BD"
THEME_TINT = 0.4  # 「Accent1, Lighter 40%」

# --- 语义色（普通 RGB 填充，颜色编码属性）---------------------------------
YELLOW = "FFFF00"  # 黄色 = 新产品线
GREEN = "92D050"   # 绿色 = 成熟产品线
GREY = "D9D9D9"    # 表头底色（无属性语义）

# 数字格式串
FMT_PCT = "0.0%"
FMT_THOUSANDS = "#,##0"
FMT_CURRENCY = '"$"#,##0.00'
FMT_GENERAL = "General"


def apply_tint(rgb_hex: str, tint: float) -> str:
    """OOXML tint 算法：对 HSL 亮度施加 tint，返回 'RRGGBB' 大写。"""
    r = int(rgb_hex[0:2], 16) / 255.0
    g = int(rgb_hex[2:4], 16) / 255.0
    b = int(rgb_hex[4:6], 16) / 255.0
    h, lum, s = colorsys.rgb_to_hls(r, g, b)
    if tint < 0:
        lum = lum * (1.0 + tint)
    else:
        lum = lum * (1.0 - tint) + (1.0 - (1.0 - tint))
    r2, g2, b2 = colorsys.hls_to_rgb(h, lum, s)
    return f"{round(r2 * 255):02X}{round(g2 * 255):02X}{round(b2 * 255):02X}"


THEME_ACCENT1_RESOLVED = apply_tint(THEME_ACCENT1_RGB, THEME_TINT)  # -> '95B3D7'


def _rgb_fill(rgb: str) -> PatternFill:
    return PatternFill(patternType="solid", fgColor=rgb)


def _theme_fill(theme_index: int, tint: float) -> PatternFill:
    return PatternFill(patternType="solid", fgColor=Color(theme=theme_index, tint=tint))


# 逐格真值累积器：cell_ref -> {期望字段}
_cell_truth: dict[str, dict] = {}


def _truth(sheet: str, ref: str, **kw) -> None:
    _cell_truth[f"{sheet}!{ref}"] = {"sheet": sheet, "cell_ref": ref, **kw}


# ---------------------------------------------------------------------------
# Sheet 1: 正常指标×期间数据区 + 语义色 + 图例 + theme/tint 填充 + 各类数字格式
# ---------------------------------------------------------------------------

def _build_data_sheet(wb: Workbook) -> tuple[str, dict]:
    sheet = "HK_Performance"
    ws = wb.create_sheet(sheet)

    # A1 实体约定（与现有 xlsx_extractor 一致）
    ws["A1"] = "ACME Hong Kong"

    # 期间表头：B1..D1
    periods = ["FY2022", "FY2023", "FY2024"]
    for j, p in enumerate(periods, start=2):
        c = ws.cell(row=1, column=j, value=p)
        c.fill = _rgb_fill(GREY)
        _truth(sheet, f"{get_column_letter(j)}1", value=p, resolved_rgb=GREY,
               number_format=FMT_GENERAL, is_merged_origin=False, merge_span=None,
               cf_affected=False, tags={})

    # A1 真值（实体标签）
    _truth(sheet, "A1", value="ACME Hong Kong", resolved_rgb=None,
           number_format=FMT_GENERAL, is_merged_origin=False, merge_span=None,
           cf_affected=False, tags={})

    # 指标行：REVENUE(新产品线,黄)/NEWSALES(成熟,绿)/PROFIT(成熟,绿)/ROE(百分比,theme底)
    # 数值硬编码
    data = {
        "REVENUE": ([2100.0, 2350.0, 2680.0], FMT_THOUSANDS, YELLOW, {"product_line": "new"}),
        "NEWSALES":  ([3800.0, 4200.0, 4750.0], FMT_THOUSANDS, GREEN, {"product_line": "mature"}),
        "PROFIT": ([1900.0, 2050.0, 2210.0], FMT_CURRENCY, GREEN, {"product_line": "mature"}),
    }
    row = 2
    for metric, (vals, fmt, rgb, tags) in data.items():
        ws.cell(row=row, column=1, value=metric)
        _truth(sheet, f"A{row}", value=metric, resolved_rgb=None,
               number_format=FMT_GENERAL, is_merged_origin=False, merge_span=None,
               cf_affected=False, tags={})
        for j, v in enumerate(vals, start=2):
            c = ws.cell(row=row, column=j, value=v)
            c.number_format = fmt
            c.fill = _rgb_fill(rgb)
            _truth(sheet, f"{get_column_letter(j)}{row}", value=v, resolved_rgb=rgb,
                   number_format=fmt, is_merged_origin=False, merge_span=None,
                   cf_affected=False, tags=dict(tags))
        row += 1

    # ROE 行：百分比格式 + theme+tint 填充（专门考验 theme 解析）
    ws.cell(row=row, column=1, value="ROE")
    _truth(sheet, f"A{row}", value="ROE", resolved_rgb=None,
           number_format=FMT_GENERAL, is_merged_origin=False, merge_span=None,
           cf_affected=False, tags={})
    roe_vals = [0.125, 0.131, 0.138]
    for j, v in enumerate(roe_vals, start=2):
        c = ws.cell(row=row, column=j, value=v)
        c.number_format = FMT_PCT
        c.fill = _theme_fill(THEME_ACCENT1_INDEX, THEME_TINT)
        _truth(sheet, f"{get_column_letter(j)}{row}", value=v,
               resolved_rgb=THEME_ACCENT1_RESOLVED, number_format=FMT_PCT,
               is_merged_origin=False, merge_span=None, cf_affected=False, tags={})

    # 图例区（右侧 F/G 列）：色块格 + 文字
    legend = [
        (YELLOW, "黄色=新产品线", "product_line", "new"),
        (GREEN, "绿色=成熟产品线", "product_line", "mature"),
    ]
    legend_expect = []
    for i, (rgb, text, tk, tv) in enumerate(legend, start=2):
        swatch_ref = f"F{i}"
        text_ref = f"G{i}"
        sc = ws[swatch_ref]
        sc.fill = _rgb_fill(rgb)
        ws[text_ref] = text
        _truth(sheet, swatch_ref, value=None, resolved_rgb=rgb,
               number_format=FMT_GENERAL, is_merged_origin=False, merge_span=None,
               cf_affected=False, tags={})
        _truth(sheet, text_ref, value=text, resolved_rgb=None,
               number_format=FMT_GENERAL, is_merged_origin=False, merge_span=None,
               cf_affected=False, tags={})
        legend_expect.append({"rgb": rgb, "meaning": text, "tag_key": tk,
                              "tag_value": tv, "swatch_ref": swatch_ref,
                              "text_ref": text_ref})

    return sheet, {
        "legend_expect": legend_expect,
        "color_clusters_expect": _expected_clusters(sheet),
    }


def _expected_clusters(sheet: str) -> list[dict]:
    """从已记录的逐格真值里推导该 sheet 的同色簇期望（跳过 cf_affected / None）。"""
    by_rgb: dict[str, list[str]] = {}
    for _key, t in _cell_truth.items():
        if t["sheet"] != sheet:
            continue
        rgb = t.get("resolved_rgb")
        if rgb is None or t.get("cf_affected"):
            continue
        by_rgb.setdefault(rgb, []).append(t["cell_ref"])
    ordered = sorted(by_rgb.items(), key=lambda item: (-len(item[1]), item[0]))
    return [{"rgb": rgb, "cell_refs": sorted(refs), "count": len(refs)}
            for rgb, refs in ordered]


# ---------------------------------------------------------------------------
# Sheet 2: 三级合并表头
# ---------------------------------------------------------------------------

def _build_merged_header_sheet(wb: Workbook) -> tuple[str, dict]:
    sheet = "MergedHeader"
    ws = wb.create_sheet(sheet)

    # 三级表头：
    #   行1: A1:F1 = "ACME Hong Kong"（一级，跨6列）
    #   行2: A2:C2 = "Agency"，D2:F2 = "Banca"（二级，各跨3列）
    #   行3: 期间 FY2022/FY2023/FY2024 × 2（三级，单格）
    merges = []

    ws.merge_cells("A1:F1")
    ws["A1"] = "ACME Hong Kong"
    _truth(sheet, "A1", value="ACME Hong Kong", resolved_rgb=None,
           number_format=FMT_GENERAL, is_merged_origin=True, merge_span=[1, 6],
           cf_affected=False, tags={})
    merges.append({"range": "A1:F1", "origin": "A1", "value": "ACME Hong Kong",
                   "span": [1, 6]})

    ws.merge_cells("A2:C2")
    ws["A2"] = "Agency"
    _truth(sheet, "A2", value="Agency", resolved_rgb=None,
           number_format=FMT_GENERAL, is_merged_origin=True, merge_span=[1, 3],
           cf_affected=False, tags={})
    merges.append({"range": "A2:C2", "origin": "A2", "value": "Agency", "span": [1, 3]})

    ws.merge_cells("D2:F2")
    ws["D2"] = "Banca"
    _truth(sheet, "D2", value="Banca", resolved_rgb=None,
           number_format=FMT_GENERAL, is_merged_origin=True, merge_span=[1, 3],
           cf_affected=False, tags={})
    merges.append({"range": "D2:F2", "origin": "D2", "value": "Banca", "span": [1, 3]})

    periods = ["FY2022", "FY2023", "FY2024", "FY2022", "FY2023", "FY2024"]
    for j, p in enumerate(periods, start=1):
        c = ws.cell(row=3, column=j, value=p)
        _truth(sheet, f"{get_column_letter(j)}3", value=p, resolved_rgb=None,
               number_format=FMT_GENERAL, is_merged_origin=False, merge_span=None,
               cf_affected=False, tags={})

    # 一行数据（REVENUE），证明合并表头下数字不张冠李戴
    revenue = [1200.0, 1320.0, 1450.0, 900.0, 1030.0, 1230.0]
    for j, v in enumerate(revenue, start=1):
        c = ws.cell(row=4, column=j, value=v)
        c.number_format = FMT_THOUSANDS
        _truth(sheet, f"{get_column_letter(j)}4", value=v, resolved_rgb=None,
               number_format=FMT_THOUSANDS, is_merged_origin=False, merge_span=None,
               cf_affected=False, tags={})

    return sheet, {"merges_expect": merges}


# ---------------------------------------------------------------------------
# Sheet 3: 转置表（指标在列、期间在行）
# ---------------------------------------------------------------------------

def _build_transposed_sheet(wb: Workbook) -> tuple[str, dict]:
    sheet = "Transposed"
    ws = wb.create_sheet(sheet)

    ws["A1"] = "Period"
    metrics = ["REVENUE", "NEWSALES", "PROFIT"]
    for j, m in enumerate(metrics, start=2):
        ws.cell(row=1, column=j, value=m)
        _truth(sheet, f"{get_column_letter(j)}1", value=m, resolved_rgb=None,
               number_format=FMT_GENERAL, is_merged_origin=False, merge_span=None,
               cf_affected=False, tags={})
    _truth(sheet, "A1", value="Period", resolved_rgb=None, number_format=FMT_GENERAL,
           is_merged_origin=False, merge_span=None, cf_affected=False, tags={})

    rows = {
        "FY2023": [2350.0, 4200.0, 2050.0],
        "FY2024": [2680.0, 4750.0, 2210.0],
    }
    for i, (period, vals) in enumerate(rows.items(), start=2):
        ws.cell(row=i, column=1, value=period)
        _truth(sheet, f"A{i}", value=period, resolved_rgb=None, number_format=FMT_GENERAL,
               is_merged_origin=False, merge_span=None, cf_affected=False, tags={})
        for j, v in enumerate(vals, start=2):
            c = ws.cell(row=i, column=j, value=v)
            c.number_format = FMT_THOUSANDS
            _truth(sheet, f"{get_column_letter(j)}{i}", value=v, resolved_rgb=None,
                   number_format=FMT_THOUSANDS, is_merged_origin=False, merge_span=None,
                   cf_affected=False, tags={})

    return sheet, {"orientation_expect": "transposed"}


# ---------------------------------------------------------------------------
# Sheet 4: 条件格式区域（色阶）—— 受影响格须打 cf_affected 并产 grid 告警
# ---------------------------------------------------------------------------

def _build_cf_sheet(wb: Workbook) -> tuple[str, dict]:
    sheet = "CondFormat"
    ws = wb.create_sheet(sheet)

    ws["A1"] = "Metric"
    ws["B1"] = "FY2024"
    _truth(sheet, "A1", value="Metric", resolved_rgb=None, number_format=FMT_GENERAL,
           is_merged_origin=False, merge_span=None, cf_affected=False, tags={})
    _truth(sheet, "B1", value="FY2024", resolved_rgb=None, number_format=FMT_GENERAL,
           is_merged_origin=False, merge_span=None, cf_affected=False, tags={})

    vals = [("REVENUE", 2680.0), ("NEWSALES", 4750.0), ("PROFIT", 2210.0), ("ROE", 0.138)]
    cf_range = "B2:B5"
    for i, (metric, v) in enumerate(vals, start=2):
        ws.cell(row=i, column=1, value=metric)
        _truth(sheet, f"A{i}", value=metric, resolved_rgb=None, number_format=FMT_GENERAL,
               is_merged_origin=False, merge_span=None, cf_affected=False, tags={})
        fmt = FMT_PCT if metric == "ROE" else FMT_THOUSANDS
        c = ws.cell(row=i, column=2, value=v)
        c.number_format = fmt
        # CF 区域内的格：resolved_rgb 不可信 -> None，cf_affected=True
        _truth(sheet, f"B{i}", value=v, resolved_rgb=None, number_format=fmt,
               is_merged_origin=False, merge_span=None, cf_affected=True, tags={})

    ws.conditional_formatting.add(
        cf_range,
        ColorScaleRule(
            start_type="min", start_color="FFF8696B",
            mid_type="percentile", mid_value=50, mid_color="FFFFEB84",
            end_type="max", end_color="FF63BE7B",
        ),
    )

    return sheet, {"cf_ranges_expect": [cf_range]}


# ---------------------------------------------------------------------------
# 组装 + 真值落盘 + 自校验
# ---------------------------------------------------------------------------

def build_workbook() -> dict:
    wb = Workbook()
    # 删掉默认空 sheet
    default = wb.active
    wb.remove(default)

    sheet1, meta1 = _build_data_sheet(wb)
    sheet2, meta2 = _build_merged_header_sheet(wb)
    sheet3, meta3 = _build_transposed_sheet(wb)
    sheet4, meta4 = _build_cf_sheet(wb)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(str(XLSX_PATH))

    ground_truth = {
        "file": XLSX_PATH.name,
        "theme_palette": {
            "accent1_index": THEME_ACCENT1_INDEX,
            "accent1_base_rgb": THEME_ACCENT1_RGB,
            "accent1_tint": THEME_TINT,
            "accent1_resolved_rgb": THEME_ACCENT1_RESOLVED,
        },
        "sheets": {
            sheet1: meta1,
            sheet2: meta2,
            sheet3: meta3,
            sheet4: meta4,
        },
        "cells": list(_cell_truth.values()),
    }
    GT_PATH.write_text(json.dumps(ground_truth, ensure_ascii=False, indent=2),
                       encoding="utf-8")
    return ground_truth


def self_verify() -> None:
    """用 openpyxl 重新读回，断言值与样式写入成功（fixture 可靠性自校验）。"""
    wb = load_workbook(str(XLSX_PATH))

    # 1) 值：抽样核对硬编码数值与文本
    ws1 = wb["HK_Performance"]
    assert ws1["A1"].value == "ACME Hong Kong", ws1["A1"].value
    assert ws1["B2"].value == 2100.0, ws1["B2"].value
    assert ws1["B5"].value == 0.125, ws1["B5"].value

    # 2) 普通 RGB 填充
    fg = ws1["B2"].fill.fgColor
    assert fg.type == "rgb" and fg.rgb.endswith(YELLOW), (fg.type, fg.rgb)

    # 3) theme + tint 填充（ROE 行）
    roe_fg = ws1["B5"].fill.fgColor
    assert roe_fg.type == "theme", roe_fg.type
    assert roe_fg.theme == THEME_ACCENT1_INDEX, roe_fg.theme
    assert abs(roe_fg.tint - THEME_TINT) < 1e-9, roe_fg.tint

    # 4) 数字格式
    assert ws1["B2"].number_format == FMT_THOUSANDS, ws1["B2"].number_format
    assert ws1["B5"].number_format == FMT_PCT, ws1["B5"].number_format
    assert ws1["B4"].number_format == FMT_CURRENCY, ws1["B4"].number_format

    # 5) 合并表头
    ws2 = wb["MergedHeader"]
    ranges = {str(r) for r in ws2.merged_cells.ranges}
    assert {"A1:F1", "A2:C2", "D2:F2"} <= ranges, ranges
    assert ws2["A1"].value == "ACME Hong Kong"

    # 6) 转置表
    ws3 = wb["Transposed"]
    assert ws3["A1"].value == "Period"
    assert ws3["B1"].value == "REVENUE"

    # 7) 条件格式
    ws4 = wb["CondFormat"]
    cf_ranges = {str(cf.sqref) for cf in ws4.conditional_formatting}
    assert "B2:B5" in cf_ranges, cf_ranges

    # 8) 图例区
    assert ws1["G2"].value == "黄色=新产品线", ws1["G2"].value
    assert ws1["F2"].fill.fgColor.rgb.endswith(YELLOW)

    print("self-verify: 值与样式 round-trip 全部通过")


def main() -> None:
    gt = build_workbook()
    self_verify()
    n_cells = len(gt["cells"])
    print(f"fixture 已生成于 {OUT_DIR}")
    print(f"  xlsx: {XLSX_PATH.name}（{len(gt['sheets'])} sheets）")
    print(f"  ground_truth: {GT_PATH.name}（{n_cells} 逐格真值）")
    print(f"  theme accent1+tint{THEME_TINT} -> {THEME_ACCENT1_RESOLVED}")


if __name__ == "__main__":
    main()
