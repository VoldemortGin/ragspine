"""单文件 Excel ingestion 编排：styled_grid 抽取 → glossary 归一 → 颜色 tags → 入库。

把一期已落地的 L1/L2 组件串成端到端入库管线，并产出可审计的 IngestReport：

    extract_grids（xlsx_styled_extractor，L1 确定性样式网格）
      → 沿用现有 xlsx 5yr 约定定位指标×期间×实体（首列指标 / 首行期间 /
        A1 或 sheet 名定实体，fixture HK_Performance 即该布局）
      → glossary 归一（metric_code / entity / geography / period / unit）
      → 颜色 tags 经 MappingRegistry 的 active 映射翻译（apply_mapping，
        未确认映射 tags 置空并告警 —— 绝不静默入库）
      → 入 fact_store，带血缘新字段（source_file_hash / extractor_version /
        mapping_version / tags / confidence / review_status）。

关键外部行为约定（PRD user stories #16/#17）：
    - dry_run=True：完整跑抽取并产出完整报告（n_grids / n_facts_extracted /
      告警 / tags 统计照常），但 store 和 queue **零写入**，n_facts_ingested 必为 0。
    - 同一文件重复 ingest 完全幂等：fact_store 唯一键 upsert，不产生重复事实
      （ingest 两次后 store.count() 不变）；报告可见幂等（重复入库 n_facts_ingested
      仍为该文件应入库的事实数，但库内总条数不增长）。
    - 抽取/入库过程中的 grid 告警与归一失败告警都汇聚进 report.warnings。
    - 任意环节抛错时 status='failed' 且 error 记录原因，已成功部分按实现的事务策略处理。

实现已完成，dataclass 字段契约保持冻结，行为契约见 tests/ingestion/structured/test_ingestion.py。
"""

from collections import Counter
from collections.abc import Sequence
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, cast

from openpyxl.utils import get_column_letter

from ragspine.common.company_profile import DimensionSpec, load_company_profile
from ragspine.common.glossary import (
    geography_for_entity,
    normalize_entity,
    normalize_metric,
    normalize_period,
    unit_for_metric,
)
from ragspine.extraction.color.color_semantics import MappingRegistry, apply_mapping
from ragspine.extraction.extractors import (
    docspine_extractor,
    pdf_digital_extractor,
    pdf_scanned_extractor,
    pdf_spine_extractor,
    pptx_styled_extractor,
)
from ragspine.extraction.extractors.xlsx_styled_extractor import extract_grids
from ragspine.extraction.ir import StyledGrid
from ragspine.extraction.registry import Extractor
from ragspine.extraction.routing import pdf_router
from ragspine.ingestion.review.review_queue import ReviewQueue
from ragspine.ingestion.structured.ingestion_manifest import ManifestStore
from ragspine.storage.fact_store import REVIEW_AUTO_APPROVED, Fact, FactStore

# home 领域 profile（声明维度默认值的来源，配置化，不硬编码）。抽取兜底的
# channel / geography / unit 默认皆从活跃 profile 的维度派生，调用期读取本全局。
_PROFILE = load_company_profile()


def _dim(name: str) -> DimensionSpec | None:
    """按名取当前 _PROFILE 的维度规格（缺失返回 None）。与 glossary/intent 同形。"""
    return next((d for d in _PROFILE.dimensions if d.name == name), None)


def _channel_default() -> str:
    """实体未细分渠道时的兜底渠道（金融默认 = 'TOTAL'，取自 channel 维 default）。"""
    dim = _dim("channel")
    return (dim.default if dim is not None else None) or "TOTAL"


def _geography_default() -> str:
    """实体未映射到地理时的兜底口径（金融默认 = 'UNKNOWN'，取自 geography 维 default）。"""
    dim = _dim("geography")
    return (dim.default if dim is not None else None) or "UNKNOWN"


def _default_unit() -> str:
    """指标无受控单位时的兜底单位：取 metric 维 units 取值的众数（金融默认 = 'USD_M'）。"""
    dim = _dim("metric")
    units = dim.units if dim is not None else {}
    if not units:
        return "USD_M"
    return Counter(units.values()).most_common(1)[0][0]


@dataclass
class IngestReport:
    """单次 ingestion 的结构化报告（dry_run 与正式入库共用）。

    字段语义约定：
        source_path:        被 ingest 的源文件路径（字符串）。
        source_doc_id:      源文件名（= fact.source_doc_id，血缘根）。
        file_hash:          源文件内容 hash（= fact.source_file_hash，版本血缘）。
        dry_run:            本次是否为 dry-run（True 时 store/queue 零写入）。
        n_grids:            抽取出的 StyledGrid 数（每 worksheet 一张）。
        n_facts_extracted:  从网格识别出的候选事实数（归一成功的指标×期间×实体）。
        n_facts_ingested:   实际写入 fact_store 的事实数；dry_run=True 时**必为 0**。
        n_tags_applied:     被 active 颜色映射成功打上 tag 的事实数。
        warnings:           汇聚的告警（grid 级告警 + 归一失败 + 未确认映射等）。
        n_enqueued_review:  本次入复核队列的项数（低置信 / 未确认映射等）。
        status:             'ok' 正常结束 / 'failed' 中途失败。
        error:              失败原因（status='ok' 时为 None）。
    """

    source_path: str
    source_doc_id: str
    file_hash: str | None = None
    dry_run: bool = False
    n_grids: int = 0
    n_facts_extracted: int = 0
    n_facts_ingested: int = 0
    n_tags_applied: int = 0
    warnings: list[str] = field(default_factory=list)
    n_enqueued_review: int = 0
    status: str = "ok"
    error: str | None = None


def _coerce_number(value: object) -> float | None:
    """单元格值转数值；非数值返回 None（沿用 MVP xlsx_extractor 约定）。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "").replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _grid_value(grid: StyledGrid, col: int, row: int) -> object:
    """取网格 (col, row) 处的原始值；空格返回 None。

    兼容两种坐标键：xlsx 用 'A1' 风格（get_column_letter），pptx/pdf 表格用
    'R{行}C{列}' 风格。先按 A1 取，未命中再按 R/C 取，使同一抽取逻辑可跨格式复用。
    """
    cell = grid.get(f"{get_column_letter(col)}{row}")
    if cell is None:
        cell = grid.get(f"R{row}C{col}")
    return cell.value if cell is not None else None


def _grid_cell_ref(grid: StyledGrid, col: int, row: int) -> str:
    """取网格 (col, row) 处单元格的原生 cell_ref（血缘 locator 用）。

    xlsx 用 'A1' 风格（cell_ref='B2'），pptx/pdf/扫描 OCR 表用 'R{行}C{列}' 风格
    （cell_ref='R2C2'）。直接回传该格自带的 cell_ref，使血缘 locator 反映其真实坐标
    （扫描 OCR 事实因此带 'page{N}_table{M}!R{r}C{c}'）；缺失时回退 A1 推导。
    """
    cell = grid.get(f"{get_column_letter(col)}{row}") or grid.get(f"R{row}C{col}")
    return cell.cell_ref if cell is not None else f"{get_column_letter(col)}{row}"


def _grid_has_candidate_data(grid: StyledGrid) -> bool:
    """该表是否含可抽取的「指标×期间×数值」候选数据（与抽取布局同套规则）。

    只在实体无法解析时用于区分「确含数据但无法归因的表」与「无数据的辅助表 /
    版式演示表」：前者入复核请人工指认实体，后者静默跳过（不滥发复核）。
    """
    periods = [
        col for col in range(2, grid.n_cols + 1)
        if (raw := _grid_value(grid, col, 1)) is not None
        and normalize_period(str(raw)) is not None
    ]
    if not periods:
        return False
    for row in range(2, grid.n_rows + 1):
        label = _grid_value(grid, 1, row)
        if label is None or normalize_metric(str(label)) is None:
            continue
        if any(_coerce_number(_grid_value(grid, col, row)) is not None for col in periods):
            return True
    return False


def _extract_facts_from_grid(
    grid: StyledGrid,
    cell_tags: dict[str, dict[str, str]],
    extractor_version: str,
) -> tuple[list[Fact], list[str], int, bool]:
    """从一张 StyledGrid 按「首列指标 / 首行期间 / A1 实体」约定抽出候选 Fact。

    复用 MVP xlsx_extractor 的确定性布局规则；颜色 tags 由 cell_tags（apply_mapping
    结果，{cell_ref: {tag_key: tag_value}}）按单元格坐标合并到对应事实上。
    返回 (facts, warnings, n_tags_applied, unattributable_data)：unattributable_data=True
    表示该表 A1/sheet 名都无法归一成受控实体、但确含可抽取数据（指标×期间×数值）
    —— 绝不臆造实体，由编排层据此入复核队列（实体无法解析，需人工指认）。
    """
    facts: list[Fact] = []
    warnings: list[str] = []
    n_tags_applied = 0

    a1 = cast("str | None", _grid_value(grid, 1, 1))
    entity = normalize_entity(a1) or normalize_entity(grid.sheet)
    if entity is None:
        warnings.append(
            f"sheet={grid.sheet}: A1/sheet 名无法解析实体，跳过该表"
        )
        # entity_unresolved=True；第 4 位仅当该表确含可抽取数据（指标×期间×数值）时
        # 报 has_candidate_data，供编排层判断是否「整文件不可归因」入复核（绝不臆造实体）。
        return facts, warnings, n_tags_applied, _grid_has_candidate_data(grid)
    geography = geography_for_entity(entity) or _geography_default()

    # 期间表头：第 1 行 B 列起。
    periods: dict[int, tuple[tuple[str, str], str]] = {}
    for col in range(2, grid.n_cols + 1):
        raw = _grid_value(grid, col, 1)
        parsed = normalize_period(str(raw)) if raw is not None else None
        if parsed is None:
            if raw is not None:
                warnings.append(
                    f"sheet={grid.sheet}!{get_column_letter(col)}1: "
                    f"无法识别期间 '{raw}'，跳过该列"
                )
            continue
        periods[col] = (parsed, str(raw).strip())

    # 指标行：第 2 行起 A 列。
    for row in range(2, grid.n_rows + 1):
        metric_label = _grid_value(grid, 1, row)
        metric_code = (
            normalize_metric(str(metric_label)) if metric_label is not None else None
        )
        if metric_code is None:
            if metric_label is not None:
                warnings.append(
                    f"sheet={grid.sheet}!A{row}: 无法识别指标 '{metric_label}'，跳过该行"
                )
            continue
        unit = unit_for_metric(metric_code) or _default_unit()
        for col, ((period_type, period), _raw) in periods.items():
            coord = f"{get_column_letter(col)}{row}"
            value = _coerce_number(_grid_value(grid, col, row))
            if value is None:
                continue
            tags: dict[str, object] = dict(cell_tags.get(coord, {}))
            if tags:
                n_tags_applied += 1
            ref = _grid_cell_ref(grid, col, row)
            facts.append(
                Fact(
                    metric_code=metric_code,
                    entity=entity,
                    geography=geography,
                    channel=_channel_default(),
                    period_type=period_type,
                    period=period,
                    value=value,
                    unit=unit,
                    source_doc_id=grid.source_doc_id,
                    source_locator=f"sheet={grid.sheet}!{ref}",
                    tags=tags,
                    source_file_hash=grid.source_file_hash,
                    extractor_version=extractor_version,
                    mapping_version=None,
                    review_status=REVIEW_AUTO_APPROVED,
                )
            )

    return facts, warnings, n_tags_applied, False


def _dedup_first(facts: list[Fact]) -> list[Fact]:
    """按事实身份键去重并保留首次出现（与 fact_store dim_key 同口径）。

    身份键 = (metric_code, entity, channel, period_type, period)；geography 非身份维
    （可被覆盖），不参与键。保持输入顺序，使「首页先于次页」时血缘锚定首页。
    """
    seen: set[tuple[str, str, str, str, str]] = set()
    out: list[Fact] = []
    for f in facts:
        key = (f.metric_code, f.entity, f.channel, f.period_type, f.period)
        if key in seen:
            continue
        seen.add(key)
        out.append(f)
    return out


def _grid_has_colored_cells(grid: StyledGrid) -> bool:
    """该网格是否含「可靠着色」的格（resolved_rgb 非空且非条件格式）。

    用于判断文件确有颜色编码需翻译成 tag —— 若此时 scope 无 active 颜色映射，
    编排层应入复核（颜色映射未确认），绝不静默把带语义的颜色当无意义忽略。
    """
    return any(cell.rgb_tag_key() is not None for cell in grid.iter_cells())


def _ingest_grids(
    report: IngestReport,
    grids: Sequence[StyledGrid],
    store: FactStore,
    registry: MappingRegistry,
    queue: ReviewQueue,
    *,
    dry_run: bool,
    extractor_version: str,
    valid_as_of: str | None = None,
    dedup_identity: bool = False,
) -> None:
    """把一组已抽取的 StyledGrid 跑完归一 → 颜色 tag → 入库 / 入队，填充 report。

    dedup_identity=True 时（扫描型多页 OCR 路径用）：入库前按事实身份键
    （metric/entity/channel/period_type/period，与 fact_store dim_key 一致）去重并
    **保留首次出现**，使同一事实跨多页时血缘锚定到首页（确定性 first-wins，而非
    upsert 的 last-wins）。xlsx/pptx/pdf-digital 路径不传该参数 -> 行为字节级不变。

    ingest_excel 与 ingest_file 的 xlsx/pptx/pdf(digital) 分支共用此逻辑，保证
    Excel 既有行为不变（resolvable 实体 + 有 active 映射 -> 零 enqueue）。

    复核入队（仅在【无法自信自动入库】时触发，绝不滥发）：
        - 整文件无任何 grid 能归一出实体、却确含可抽取数据 -> 该（些）表跳过 +
          enqueue（reason 含「实体无法解析」），绝不臆造实体造成 ingestion 误归因；
          若文件中至少一张表的实体可解析（如多 sheet 工作簿的主数据表），则其余
          无法归因的辅助 / 版式表沿用既有「告警跳过」语义，不滥发复核。
        - 文件确有颜色 tag 需翻译但 scope 无 active 颜色映射 -> enqueue（颜色映射未确认）。
    dry_run=True 时 store 与 queue 都零写入（n_facts_ingested / n_enqueued_review 皆 0）。
    """
    report.n_grids = len(grids)
    if grids:
        report.file_hash = grids[0].source_file_hash

    source_doc_id = report.source_doc_id

    # active 颜色映射（scope = 源文件名）；无 active 映射 -> 告警、tags 置空。
    active_mapping = registry.get_active(source_doc_id)
    if active_mapping is None:
        report.warnings.append(
            f"scope={source_doc_id} 无 active 颜色映射（color mapping 未确认），"
            "相关颜色 tags 置空"
        )

    all_facts: list[Fact] = []
    needs_color_mapping = False
    enqueues: list[tuple[str, dict[str, Any], str]] = []  # (reason, payload, locator)
    any_entity_resolved = False
    unattributable_grids: list[StyledGrid] = []  # 实体不可解析但确含数据的表（待文件级裁决）
    for grid in grids:
        # grid 级告警（条件格式 / 来源不可靠等）汇聚进报告。
        report.warnings.extend(grid.warnings)
        cell_tags = (
            apply_mapping(grid, active_mapping) if active_mapping is not None else {}
        )
        facts, warnings, n_tags, unattributable = _extract_facts_from_grid(
            grid, cell_tags, extractor_version
        )
        all_facts.extend(facts)
        report.warnings.extend(warnings)
        report.n_tags_applied += n_tags

        if facts:
            any_entity_resolved = True
        if unattributable:
            unattributable_grids.append(grid)
        if active_mapping is None and _grid_has_colored_cells(grid):
            needs_color_mapping = True

    # 文件级裁决：仅当【整文件无任何表归因成功】时，才把含数据的不可归因表入复核
    # （绝不臆造实体；多 sheet 工作簿有主数据表时，辅助表按既有语义静默跳过）。
    if not any_entity_resolved:
        for grid in unattributable_grids:
            enqueues.append((
                "实体无法解析，需人工指认",
                {
                    "source_doc_id": grid.source_doc_id,
                    "sheet": grid.sheet,
                    "file_hash": grid.source_file_hash,
                },
                f"{grid.source_doc_id}!{grid.sheet}",
            ))

    if needs_color_mapping:
        report.warnings.append(
            f"scope={source_doc_id} 颜色映射未确认但文件含颜色编码，颜色 tag 未翻译，入复核"
        )
        enqueues.append((
            "颜色映射未确认，需 SME 确认图例",
            {"source_doc_id": source_doc_id, "file_hash": report.file_hash},
            source_doc_id,
        ))

    report.n_facts_extracted = len(all_facts)

    if dry_run:
        # 零写入：报告完整但不碰 store / queue。
        report.n_facts_ingested = 0
        report.n_enqueued_review = 0
        return

    # 扫描多页 OCR：按身份键去重保留首次出现（first-wins，血缘锚定首页）。
    facts_to_write = _dedup_first(all_facts) if dedup_identity else all_facts

    # 本批生效日（CLI / 调用方提供）注入每条 fact 的 valid_as_of；None 时不改写。
    if valid_as_of is not None:
        for f in facts_to_write:
            f.valid_as_of = valid_as_of
    n_written = store.upsert_facts(facts_to_write)
    report.n_facts_ingested = n_written

    for reason, payload, locator in enqueues:
        queue.enqueue(reason, payload, locator)
    report.n_enqueued_review = len(enqueues)


def ingest_excel(
    path: str | Path,
    store: FactStore,
    registry: MappingRegistry,
    queue: ReviewQueue,
    *,
    dry_run: bool = False,
    extractor_version: str = "xlsx_styled@1",
    manifest: ManifestStore | None = None,
    batch_id: str | None = None,
) -> IngestReport:
    """编排单个 xlsx 的端到端 ingestion，返回 IngestReport。

    参数：
        path:              源 xlsx 路径。
        store:             FactStore 实例（已 init_schema）。
        registry:          MappingRegistry 实例（颜色映射注册表，取 active 映射打 tag）。
        queue:             ReviewQueue 实例（低置信 / 未确认映射入复核）。
        dry_run:           True 时完整跑抽取并产出报告，但 store/queue 零写入
                           （report.n_facts_ingested == 0）。
        extractor_version: 写入每条 fact.extractor_version 的版本串。
        manifest:          可选 ManifestStore（传入则把本文件登记进当前批次台账）。
        batch_id:          可选批次 id（与 manifest 配合，关联到某一批）。

    行为见模块 docstring：抽取 → 归一 → active 映射打 tag → 入库（带血缘），
    dry_run 零写入，重复 ingest 幂等，告警汇聚，失败时 status='failed'+error。
    """
    path = Path(path)
    source_doc_id = path.name
    report = IngestReport(
        source_path=str(path),
        source_doc_id=source_doc_id,
        dry_run=dry_run,
    )

    try:
        grids = extract_grids(path)
    except Exception as exc:  # noqa: BLE001 —— 失败不裸抛，落进报告（story #16）
        report.status = "failed"
        report.error = f"{type(exc).__name__}: {exc}"
        if manifest is not None and batch_id is not None:
            manifest.record_input(
                batch_id, str(path), None, "xlsx", failed=True, error=report.error
            )
        return report

    _ingest_grids(
        report, grids, store, registry, queue,
        dry_run=dry_run, extractor_version=extractor_version,
    )

    if dry_run:
        return report

    if manifest is not None and batch_id is not None:
        manifest.record_input(
            batch_id,
            str(path),
            report.file_hash,
            "xlsx",
            n_facts=report.n_facts_ingested,
            n_warnings=len(report.warnings),
        )

    return report


# 后缀 -> (抽取器模块, extractor_version, manifest file_type)。
_EXTRACTOR_BY_SUFFIX = {
    ".xlsx": (extract_grids, "xlsx_styled@1", "xlsx"),
    ".xlsm": (extract_grids, "xlsx_styled@1", "xlsx"),
    ".pptx": (pptx_styled_extractor.extract_grids, "pptx_styled@1", "pptx"),
    # W3b：.docx/.docm（OOXML Word）-> docspine（纯 Rust 强表 gridSpan/vMerge/嵌套）。
    # 旧版二进制 .doc（OLE/CFB）docspine 不解析，故不登记 -> 命中「不支持的文件格式」。
    ".docx": (docspine_extractor.extract_grids, docspine_extractor.EXTRACTOR_VERSION, "docx"),
    ".docm": (docspine_extractor.extract_grids, docspine_extractor.EXTRACTOR_VERSION, "docx"),
}


def ingest_file(
    path: str | Path,
    store: FactStore,
    registry: MappingRegistry,
    queue: ReviewQueue,
    *,
    dry_run: bool = False,
    manifest: ManifestStore | None = None,
    batch_id: str | None = None,
    valid_as_of: str | None = None,
    grid_extractor: pdf_digital_extractor.GridExtractor | None = None,
    ocr_backend: pdf_scanned_extractor.OcrBackend | None = None,
    pptx_extractor: Extractor | None = None,
) -> IngestReport:
    """统一多格式分发入口：按后缀把文件路由到对应抽取器并复用共享入库逻辑。

    分发约定：
        - .xlsx/.xlsm -> xlsx_styled_extractor（extractor_version=xlsx_styled@1）；
        - .pptx       -> 默认 python-pptx 的 pptx_styled_extractor（pptx_styled@1，保
          color/chart/note）；注入 pptx_extractor（如家族 PptspineGridExtractor，W3c）则
          改用它解析（更富表合并），血缘 extractor_version 随之而变（additive opt-in，
          默认不注入时字节级不变）；
        - .docx/.docm -> docspine_extractor（docspine@1，纯 Rust 强表 gridSpan/vMerge/嵌套，
          W3b）；正文段落归叙事通路（narrative_extract.extract_docx_narrative）；
        - .pdf        -> 先 pdf_router.route()：
            * digital   -> pdf_digital_extractor（pdf_digital@1）确定性入库；
            * scanned/ocr_scan/mixed -> 调家族 OCR（pdf_scanned_extractor.extract_grids，
              默认 PdfSpineOcrBackend，离线确定性 PP-OCRv5）抽表入结构化通路，低置信
              格仍进复核（W3a）；OCR 未识别出表则告警 + enqueue（绝不静默丢弃）；
            * ask_for_pptx（PowerPoint/Keynote 导出）-> warning + enqueue 请求 pptx 源
              （PRD「原生优先」，不在退化 PDF 上做不可靠抽取）；
        - 其它后缀 -> status='failed' + error（不裸抛，不污染库）。

    抽取出的 grids 全部复用 _ingest_grids（归一 / active 颜色映射 / 实体不可解析与
    颜色映射未确认入复核 / dry_run 零写入 / 幂等 upsert），保证与 ingest_excel 等价。
    向后兼容：ingest_excel 保留不变。

    valid_as_of：本批事实的「截至 / 生效」业务日期（CLI 提供，可空）；非 None 时注入
    每条入库 fact 的 valid_as_of，None 时不改写（既有行为字节级不变）。
    """
    path = Path(path)
    source_doc_id = path.name
    suffix = path.suffix.lower()
    report = IngestReport(
        source_path=str(path),
        source_doc_id=source_doc_id,
        dry_run=dry_run,
    )

    if suffix == ".pdf":
        _ingest_pdf(
            report, path, store, registry, queue,
            dry_run=dry_run, valid_as_of=valid_as_of,
            grid_extractor=grid_extractor,
            ocr_backend=ocr_backend,
        )
        if not dry_run:
            _record_manifest(
                report, path, manifest, batch_id, "pdf",
                failed=report.status == "failed",
            )
        return report

    spec = _EXTRACTOR_BY_SUFFIX.get(suffix)
    if spec is None:
        report.status = "failed"
        report.error = f"不支持的文件格式: {suffix or '(无后缀)'}"
        _record_manifest(
            report, path, manifest, batch_id, suffix.lstrip(".") or "?", failed=True
        )
        return report

    extractor, extractor_version, file_type = spec
    # W3c：.pptx 默认 python-pptx；注入 pptx_extractor 可切到更富表合并的家族 pptspine
    # （opt-in，不注入时默认分发字节级不变）。血缘 extractor_version 随注入的解析器而变。
    if suffix == ".pptx" and pptx_extractor is not None:
        extractor = pptx_extractor.extract
        extractor_version = getattr(pptx_extractor, "version", extractor_version)
    try:
        grids = extractor(path)
    except Exception as exc:  # noqa: BLE001 —— 失败不裸抛，落进报告（story #16）
        report.status = "failed"
        report.error = f"{type(exc).__name__}: {exc}"
        _record_manifest(report, path, manifest, batch_id, file_type, failed=True)
        return report

    _ingest_grids(
        report, grids, store, registry, queue,
        dry_run=dry_run, extractor_version=extractor_version,
        valid_as_of=valid_as_of,
    )
    if not dry_run:
        _record_manifest(report, path, manifest, batch_id, file_type)
    return report


def _ingest_pdf(
    report: IngestReport,
    path: Path,
    store: FactStore,
    registry: MappingRegistry,
    queue: ReviewQueue,
    *,
    dry_run: bool,
    valid_as_of: str | None = None,
    grid_extractor: pdf_digital_extractor.GridExtractor | None = None,
    ocr_backend: pdf_scanned_extractor.OcrBackend | None = None,
) -> None:
    """PDF 分支：先分诊路由，数字型走确定性表格抽取，扫描型走家族 OCR（W3a）。

    grid_extractor：数字型表格抽取器（依赖注入，缺省 = PdfSpineGridExtractor）。默认走
    pdfspine（纯 Rust、确定性、无 torch 重依赖）；换别的解析器（如 DoclingGridExtractor 兜底）
    只需注入另一个 GridExtractor 实现，血缘里的 extractor_version 随之而变。

    ocr_backend：扫描型 OCR 后端（依赖注入，缺省 = PdfSpineOcrBackend，家族离线确定性
    PP-OCRv5）。scanned/ocr_scan/mixed 判定不再只 enqueue，而是实际调用
    pdf_scanned_extractor.extract_grids 抽表入结构化通路；血缘 extractor_version 随后端
    version 而定（与 GridExtractor 同范式）。
    """
    extractor = grid_extractor or pdf_spine_extractor.PdfSpineGridExtractor()
    decision = pdf_router.route(str(path))
    report.file_hash = decision.file_hash

    if decision.error is not None:
        report.status = "failed"
        report.error = f"PDF 不可读: {decision.error}"
        return

    # PowerPoint/Keynote 导出的 PDF：原生优先，不强抽，告警 + 请 pptx 源。
    if decision.ask_for_pptx:
        report.warnings.append(
            f"{report.source_doc_id} 疑似 PowerPoint/Keynote 导出 PDF，"
            "原生优先：请提供 pptx 源后再入库"
        )
        _enqueue_or_count(
            report, queue, dry_run,
            "疑似PowerPoint导出PDF，请提供pptx源",
            {"source_doc_id": report.source_doc_id, "file_hash": report.file_hash,
             "origin_meta": decision.origin_meta},
            report.source_doc_id,
        )
        return

    if decision.verdict == pdf_router.VERDICT_DIGITAL:
        try:
            grids = extractor.extract_grids(path)
        except Exception as exc:  # noqa: BLE001
            report.status = "failed"
            report.error = f"{type(exc).__name__}: {exc}"
            return
        _ingest_grids(
            report, grids, store, registry, queue,
            dry_run=dry_run, extractor_version=extractor.version,
            valid_as_of=valid_as_of,
        )
        return

    # scanned / ocr_scan / mixed：实际调用家族 OCR 抽表入结构化通路（W3a），不再只
    # enqueue。低置信格由 extract_grids 在传入 queue 时按 'low_confidence_ocr' 入复核
    # （dry_run 不传 queue -> 零写入）。
    ocr = ocr_backend or pdf_scanned_extractor.PdfSpineOcrBackend()
    try:
        grids = pdf_scanned_extractor.extract_grids(
            path, ocr, queue=None if dry_run else queue
        )
    except Exception as exc:  # noqa: BLE001 —— 失败不裸抛，落进报告
        report.status = "failed"
        report.error = f"{type(exc).__name__}: {exc}"
        return

    # OCR 未识别出任何表格：绝不静默丢弃，留告警 + 复核项让运营看到这份文件被挡下。
    if not grids:
        report.warnings.append(
            f"{report.source_doc_id} 分诊为 {decision.verdict}，OCR 未识别出任何表格，入复核"
        )
        _enqueue_or_count(
            report, queue, dry_run,
            "扫描型PDF OCR未识别出表格，需人工复核",
            {"source_doc_id": report.source_doc_id, "file_hash": report.file_hash,
             "verdict": decision.verdict},
            report.source_doc_id,
        )
        return

    _ingest_grids(
        report, grids, store, registry, queue,
        dry_run=dry_run,
        extractor_version=getattr(ocr, "version", pdf_scanned_extractor.EXTRACTOR_VERSION),
        valid_as_of=valid_as_of,
        dedup_identity=True,
    )


def _enqueue_or_count(
    report: IngestReport,
    queue: ReviewQueue,
    dry_run: bool,
    reason: str,
    payload: dict[str, Any],
    locator: str,
) -> None:
    """统一入队入口：dry_run 时只算报告不写队列（沿用「只看不动」硬约定）。"""
    if dry_run:
        return
    queue.enqueue(reason, payload, locator)
    report.n_enqueued_review += 1


def _record_manifest(
    report: IngestReport,
    path: Path,
    manifest: ManifestStore | None,
    batch_id: str | None,
    file_type: str,
    *,
    failed: bool = False,
) -> None:
    """按真实格式把本文件登记进批次台账（manifest/batch_id 任一缺失则跳过）。"""
    if manifest is None or batch_id is None:
        return
    if failed:
        manifest.record_input(
            batch_id, str(path), None, file_type, failed=True, error=report.error
        )
        return
    manifest.record_input(
        batch_id,
        str(path),
        report.file_hash,
        file_type,
        n_facts=report.n_facts_ingested,
        n_warnings=len(report.warnings),
    )
